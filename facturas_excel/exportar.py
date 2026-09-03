"""Genera el archivo .xlsx en la maqueta exacta que espera el gestor fiscal,
colocando cada campo en la columna que indica la configuracion (ConfigColumnas).
"""

from __future__ import annotations

from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

from .config_columnas import ConfigColumnas, ETIQUETA_CABECERA
from .modelo import CAMPOS_IMPORTE, CAMPOS_PORCENTAJE, Factura

# Modo en que se escriben los numeros en las celdas:
#   "texto"  -> "1234,56" como texto (separador decimal coma)
#   "numero" -> 1234.56 como numero real de Excel
MODO_TEXTO = "texto"
MODO_NUMERO = "numero"

# Aplifisa rechaza el apunte si el nombre de la cuenta pasa de 40 caracteres
# ("El tamano del nombre excede el maximo").
MAX_NOMBRE = 40


# Palabras que no deben quedar al final de un nombre recortado.
_CONECTORES = {"DE", "DEL", "LA", "EL", "LOS", "LAS", "Y", "E"}


def _recortar_nombre(nombre: str) -> str:
    """Recorta a MAX_NOMBRE sin dejar restos feos: corta por palabra entera si
    apenas cuesta caracteres ('... IBERIA, S' -> '... IBERIA') y si no corta en
    seco, que conserva mas informacion para identificar la cuenta."""
    nombre = " ".join(str(nombre).split())
    if len(nombre) <= MAX_NOMBRE:
        return nombre
    corte = nombre[:MAX_NOMBRE]
    if not nombre[MAX_NOMBRE].isspace():
        hueco = corte.rfind(" ")
        # Retroceder solo si se pierden pocas letras; si no, corte seco.
        if hueco >= MAX_NOMBRE - 6:
            corte = corte[:hueco]
    corte = corte.rstrip(" ,.-")
    palabras = corte.split(" ")
    if len(palabras) > 1 and palabras[-1].upper() in _CONECTORES:
        corte = " ".join(palabras[:-1])
    return corte.rstrip(" ,.-")


def _fmt_importe_texto(valor: float) -> str:
    return f"{valor:.2f}".replace(".", ",")


def _fmt_porcentaje_texto(valor: float) -> str:
    # 21.0 -> "21" ; 21.5 -> "21,5"
    if float(valor).is_integer():
        return str(int(valor))
    return f"{valor:g}".replace(".", ",")


def _valor_celda(campo: str, valor, modo: str):
    """Devuelve el valor listo para escribir en la celda segun el tipo de campo."""
    if valor is None or valor == "":
        return None

    # Aplifisa: el NIF de la cuenta no puede llevar puntos, espacios ni guiones.
    if campo == "nif":
        return str(valor).strip().upper().replace(".", "").replace(" ", "").replace("-", "")

    if campo == "nombre":
        return _recortar_nombre(valor)

    es_numerico = campo in CAMPOS_IMPORTE or campo in CAMPOS_PORCENTAJE
    if not es_numerico:
        return str(valor)

    valor = float(valor)
    if modo == MODO_NUMERO:
        return round(valor, 2)

    # modo texto
    if campo in CAMPOS_PORCENTAJE:
        return _fmt_porcentaje_texto(valor)
    return _fmt_importe_texto(valor)


def exportar_excel(
    facturas: List[Factura],
    config: ConfigColumnas,
    ruta_salida: str,
    modo_numeros: str = MODO_TEXTO,
) -> str:
    """Escribe el .xlsx y devuelve la ruta.

    - Fila 1: cabecera (si config.incluye_cabecera).
    - Datos desde config.primera_fila.
    - Cada campo va a la columna (letra) que diga config.columnas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Cabecera
    if config.incluye_cabecera:
        for campo, letra in config.columnas.items():
            col = column_index_from_string(letra)
            ws.cell(row=1, column=col, value=ETIQUETA_CABECERA.get(campo, campo))

    # Datos
    fila = config.primera_fila
    for factura in facturas:
        datos = factura.campos_dict()
        for campo, letra in config.columnas.items():
            valor = _valor_celda(campo, datos.get(campo), modo_numeros)
            if valor is not None:
                col = column_index_from_string(letra)
                celda = ws.cell(row=fila, column=col, value=valor)
                # Forzar texto cuando toca, para que Excel no reinterprete la coma
                if modo_numeros == MODO_TEXTO and (
                    campo in CAMPOS_IMPORTE or campo in CAMPOS_PORCENTAJE
                ):
                    celda.number_format = "@"
        fila += 1

    wb.save(ruta_salida)
    return ruta_salida


def verificar_excel(
    facturas: List[Factura],
    config: ConfigColumnas,
    ruta: str,
    modo_numeros: str = MODO_TEXTO,
) -> List[str]:
    """Vuelve a LEER el Excel escrito y lo compara con lo que hay en pantalla.

    Es el ultimo paso antes de que los apuntes entren en Aplifisa, y hasta
    ahora era el unico que nadie comprobaba: si una columna se escribiera
    corrida, un importe se perdiera o faltara una linea, no se enteraba nadie
    hasta tener el problema dentro de la contabilidad.

    Devuelve la lista de diferencias (vacia = el archivo dice exactamente lo
    mismo que la tabla).
    """
    from openpyxl import load_workbook

    problemas: List[str] = []
    try:
        hoja = load_workbook(ruta, data_only=True).active
    except Exception as e:  # archivo abierto en Excel, disco lleno...
        return [f"No se pudo volver a abrir el archivo para comprobarlo: {e}"]

    escritas = hoja.max_row - config.primera_fila + 1
    if escritas != len(facturas):
        problemas.append(
            f"El archivo tiene {max(escritas, 0)} línea(s) y deberían ser "
            f"{len(facturas)}.")

    for i, factura in enumerate(facturas):
        fila = config.primera_fila + i
        datos = factura.campos_dict()
        for campo, letra in config.columnas.items():
            esperado = _valor_celda(campo, datos.get(campo), modo_numeros)
            escrito = hoja[f"{letra}{fila}"].value
            if esperado is None and (escrito is None or escrito == ""):
                continue
            if str(escrito) != str(esperado):
                problemas.append(
                    f"Línea {i + 1} ({factura.num_factura or 'sin nº'}), "
                    f"{ETIQUETA_CABECERA.get(campo, campo)}: el archivo pone "
                    f"«{escrito}» y debería poner «{esperado}».")
    return problemas


def totales_del_excel(config: ConfigColumnas, ruta: str) -> dict:
    """Suma los importes TAL COMO HAN QUEDADO en el archivo.

    Sirve para el contraste final: estos totales tienen que ser los mismos que
    los del resumen en pantalla.
    """
    from openpyxl import load_workbook

    suma = {"lineas": 0, "base_iva": 0.0, "cuota_iva": 0.0,
            "cuota_requiv": 0.0, "cuota_irpf": 0.0}
    try:
        hoja = load_workbook(ruta, data_only=True).active
    except Exception:
        return suma
    fila = config.primera_fila
    while fila <= hoja.max_row:
        vacia = True
        for campo in list(suma)[1:]:
            letra = config.columnas.get(campo)
            if not letra:
                continue
            valor = hoja[f"{letra}{fila}"].value
            if valor in (None, ""):
                continue
            vacia = False
            try:
                suma[campo] += float(str(valor).replace(".", "").replace(",", "."))
            except ValueError:
                pass
        if not vacia or hoja[f"{config.columnas.get('fecha', 'B')}{fila}"].value:
            suma["lineas"] += 1
        fila += 1
    for campo in list(suma)[1:]:
        suma[campo] = round(suma[campo], 2)
    return suma


def ordenar_para_exportar(facturas: List[Factura], orden: str) -> List[Factura]:
    """Ordena los apuntes para el Excel.

    Aplifisa numera las facturas recibidas SEGUN ENTRAN, asi que el orden del
    archivo decide con que numero queda registrada cada factura:

    - "pdf": tal como estan (el orden del escaneo). El apunte nº 3 es la hoja 3.
      Es lo que hace falta en un requerimiento.
    - "fecha": de la mas antigua a la mas reciente, para el registro trimestral.

    Las lineas de una misma factura (varios tipos de IVA, o el suplido) NO se
    separan nunca: se ordenan juntas, por la primera.
    """
    if orden != "fecha":
        return list(facturas)

    from .validacion import fecha_de

    grupos: List[List[Factura]] = []
    indice = {}
    for f in facturas:
        clave = ((f.num_factura or "").strip().upper(), (f.nif or "").strip().upper())
        if clave in indice and grupos[indice[clave]]:
            grupos[indice[clave]].append(f)
        else:
            indice[clave] = len(grupos)
            grupos.append([f])

    def orden_grupo(par):
        i, grupo = par
        dia = fecha_de(grupo[0].fecha)
        # Sin fecha entendible: al final, sin cambiar su orden entre ellas.
        return (dia is None, dia or date.max, i)

    ordenados = sorted(enumerate(grupos), key=orden_grupo)
    return [f for _, grupo in ordenados for f in grupo]
